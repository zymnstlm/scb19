# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   lesson_7.py
# @Software :   PyCharm
# @Time :   2021/3/4 9:34
# @company  :   湖南省零檬信息技术有限公司


'''
1、编写接口自动化的测试用例，读取excel里的测试数据  -- read_data()
2、发送接口请求，得到响应结果   -- api_func()
3、预期结果   vs    实际结果   -- done
4、写入最终的结果到excel     -- write_result()
'''

import requests
import openpyxl
# 读取数据
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 获取sheet表单
    max_row = sheet.max_row   # 获得excel最大的行数
    cases_list = []  # 定义一个空的列表用来存放测试用例
    for i in range(2, max_row+1):   # +1 是因为取头不取尾
        dict_1 = dict(
        case_id = sheet.cell(row=i, column=1).value,  # 用例编号
        url = sheet.cell(row=i, column=5).value,  # 接口地址
        data = sheet.cell(row=i, column=6).value,  # 请求参数
        excepted = sheet.cell(row=i, column=7).value)  # 预期结果
        cases_list.append(dict_1)  # 把字典一条一条的追加到列表里存储
    # print(cases_list)
    return cases_list

# 发送请求
def api_func(url, data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url,json=data,headers=header)
    res_r = res.json()
    return res_r

# 写入结果
def write_result(filename, sheetname, row, column, final_res):
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 获取sheet表单
    sheet.cell(row=row, column=column).value = final_res  # 写入结果
    wb.save(filename)   # 保存测试用例

# 执行自动化脚本
def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)  #  调用读取函数
    for case in cases:  # 从读取函数里返回的数据进行取值
        case_id = case['case_id']   # 取出用例编号
        url = case.get('url')   # 取出接口地址
        data = case.get('data') # 取出请求参数
        excepted = case.get('excepted') # 取出预期结果
        data = eval(data)   # 使用eval()函数进行类型转换 -->运行被字符串包裹的python表达式
        excepted = eval(excepted)   # 类型转换
        excepted_msg = excepted.get('msg')  # 取出预期结果里的msg去做结果判断
        real_res = api_func(url=url, data=data)    # 调用发送请求的接口
        real_res_msg = real_res.get('msg')  # 取出实际结果里的msg去做结果判断
        print('实际结果为：{}'.format(real_res_msg))
        print('预期结果为：{}'.format(excepted_msg))
        if excepted_msg == real_res_msg:
            print('这条测试用例通过!!')
            final_res = 'pass'  # 用变量来接收最终结果
        else:
            print('这条测试用例不通过！！！')
            final_res = 'NG'    # 用变量来接收最终结果
        print('*' * 40)
        write_result(filename, sheetname, case_id+1, 8, final_res) #调用回写函数

execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')




